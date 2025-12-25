import os

import openpyxl

def get_data(sheetName):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, "..", "Excel", "testdata.xlsx")
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows + 1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList
